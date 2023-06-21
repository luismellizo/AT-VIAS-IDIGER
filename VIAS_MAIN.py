from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
import pyautogui
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import openpyxl
from fpdf import FPDF
import os
import img2pdf
from openpyxl import load_workbook
import datetime
from PIL import Image
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from pywinauto.application import Application
from selenium.common.exceptions import NoSuchElementException
import pdfkit
import datetime


def convertir_imagen_a_pdf(ruta_imagen, ruta_pdf):
    c = canvas.Canvas(ruta_pdf, pagesize=letter)
    c.drawImage(ruta_imagen, 0, 0, width=letter[0], height=letter[1])
    c.save()


# NOTAS #NOTAS # NOTAS # Ruta del chromedriver.exe
driver_path = r"C:\Users\lamel\OneDrive\Documentos\PROGRAMAS PYTHON\SELENIUM\chromedriver.exe"

# NOTAS #NOTAS # NOTAS # Ruta del archivo Excel
excel_file = r"C:\Users\lamel\OneDrive\Escritorio\PRUEBA VIAS\PRUEBITA.xlsx"

# NOTAS #NOTAS # NOTAS # Ruta de la carpeta para guardar los archivos PDF
output_folder = r"D:\SERVICIOS ESPECIALES 2022\BOGOTA CONCEPTOS\VIAS OK\OFICIOS 4 JUNIO\Radicado"

# NOTAS #NOTAS # NOTAS # Inicializar el driver de Selenium
service = Service(driver_path)
driver = webdriver.Chrome(service=service)

# NOTAS #NOTAS # NOTAS # Abrir el archivo Excel
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# NOTAS #NOTAS # NOTAS # Obtener el número de filas en el archivo Excel
max_row = sheet.max_row


# NOTAS #NOTAS # NOTAS # Abrir la página web
driver.get("http://sipawebfile.sdp.gov.co/webfile/")

# NOTAS #NOTAS # NOTAS # Esperar a que la página se cargue completamente
time.sleep(3)

# NOTAS #NOTAS # NOTAS # Hacer clic en el enlace con el título "RADICACIÓN VIRTUAL"
tramite_button = driver.find_element(
    "xpath", "//a[contains(@title, 'RADICACIÓN VIRTUAL')]")
tramite_button.click()

# NOTAS #NOTAS # NOTAS # Esperar a que se abra el popup (class="popup")
time.sleep(3)

# NOTAS #NOTAS # NOTAS # Escribir en el campo de usuario
usuario_input = driver.find_element("name", "username")
usuario_input.send_keys("1013689567")

# NOTAS #NOTAS # NOTAS # Esperar 3 segundos antes de interactuar con el campo de contraseña
# NOTAS #NOTAS # NOTAS # Presionar la tecla "Tab"
usuario_input.send_keys(Keys.TAB)
# NOTAS #NOTAS # NOTAS # Escribir la contraseña en el campo activo (contraseña)
contraseña_input = driver.switch_to.active_element
contraseña_input.send_keys("Bf4HbrvnQWLr")

# NOTAS #NOTAS # NOTAS # Presionar la tecla "Tab" cuatro veces
for _ in range(4):
    ActionChains(driver).send_keys(Keys.TAB).perform()

# NOTAS #NOTAS # NOTAS # Presionar la tecla "Enter"
ActionChains(driver).send_keys(Keys.ENTER).perform()

# NOTAS #NOTAS # NOTAS # Esperar a que aparezca el elemento con id="contenido"
time.sleep(1)

# NOTAS #NOTAS # NOTAS # Presionar la tecla "Tab" 22 veces
for _ in range(22):
    ActionChains(driver).send_keys(Keys.TAB).perform()

# NOTAS #NOTAS # NOTAS # Esperar 1 segundo antes de escribir "hola"
time.sleep(1)

# NOTAS #NOTAS # NOTAS # Escribir "hola" en el campo activo
ActionChains(driver).send_keys(
    "REFERENCIA: DERECHO DE PETCIÓN CONCEPTO DE AFECTACIÓN RESERVA VIAL").perform()

# NOTAS #NOTAS # NOTAS # Esperar 1 segundo antes de escribir "hola"
time.sleep(1)

# NOTAS #NOTAS # NOTAS # Presionar la tecla "Tab" una vez
ActionChains(driver).send_keys(Keys.TAB).perform()

# NOTAS #NOTAS # NOTAS # Esperar 1 segundo antes de escribir "aquí estoy"
time.sleep(1)

# NOTAS #NOTAS # NOTAS # Escribir "aquí estoy" en el campo activo
ActionChains(driver).send_keys("PETICIÓN: Se solicita a SECRETARIA DISTRITAL DE PLANEACIÓN-SDP, concepto en el que se indique si el predio en mención presenta afectación Vial y si esta presenta restricciones para la instalación, localización y regularización de estaciones radioeléctricas. ").perform()

# NOTAS #NOTAS # NOTAS # Presionar la tecla "Tab" una vez
ActionChains(driver).send_keys(Keys.TAB).perform()

# NOTAS #NOTAS # NOTAS # Presionar la tecla "Enter"
ActionChains(driver).send_keys(Keys.ENTER).perform()

# NOTAS #NOTAS # NOTAS # Presionar la tecla "Tab" cuatro veces
for _ in range(4):
    ActionChains(driver).send_keys(Keys.TAB).perform()

# NOTAS #NOTAS # NOTAS # Presionar la tecla "Enter"
ActionChains(driver).send_keys(Keys.ENTER).perform()

# NOTAS #NOTAS # NOTAS # Presionar la tecla "Tab" una vez
ActionChains(driver).send_keys(Keys.TAB).perform()

# NOTAS #NOTAS # NOTAS # Presionar la tecla "Enter"
ActionChains(driver).send_keys(Keys.ENTER).perform()

# NOTAS #NOTAS # NOTAS # Esperar a que aparezca la ventana de selección de archivo (ventana de Windows)
time.sleep(1)

# NOTAS #NOTAS # NOTAS # Obtener la ruta del PDF de la columna B
pdf_path = sheet.cell(row=2, column=2).value

# NOTAS #NOTAS # NOTAS # Enviar la ruta del archivo al cuadro de diálogo de selección de archivo
pyautogui.write(pdf_path)

# NOTAS #NOTAS # NOTAS # Presionar la tecla "Enter" para confirmar la selección del archivo
pyautogui.press("enter")

# NOTAS #NOTAS # NOTAS # Esperar a que se cargue el archivo
time.sleep(1)

# NOTAS #NOTAS # NOTAS # Presionar la tecla "Enter" para confirmar la selección del archivo
pyautogui.press("enter")

# NOTAS #NOTAS # NOTAS # Esperar a que se cargue el archivo
time.sleep(1)

# NOTAS #NOTAS # NOTAS # Presionar la tecla "Tab" una vez
ActionChains(driver).send_keys(Keys.TAB).perform()

# NOTAS #NOTAS # NOTAS # Presionar la tecla "Enter"
ActionChains(driver).send_keys(Keys.ENTER).perform()

# NOTAS #NOTAS # NOTAS # Esperar a que se cargue el archivo
time.sleep(3)

# NOTAS #NOTAS # NOTAS # Presionar la tecla "Tab" 24 veces
for _ in range(24):
    ActionChains(driver).send_keys(Keys.TAB).perform()

# NOTAS #NOTAS # NOTAS # Presionar la tecla "Enter"
ActionChains(driver).send_keys(Keys.ENTER).perform()

# NOTAS #NOTAS # NOTAS # Presionar la tecla "Tab" 24 veces
for _ in range(30):
    ActionChains(driver).send_keys(Keys.TAB).perform()

# NOTAS #NOTAS # NOTAS # Presionar la tecla de flecha derecha
ActionChains(driver).send_keys(Keys.ARROW_RIGHT).perform()

# NOTAS #NOTAS # NOTAS # Obtener todos los iframes presentes en la página
iframes = driver.find_elements(By.TAG_NAME, "iframe")

# NOTAS #NOTAS # NOTAS # Variable para almacenar si se encuentra el elemento
elemento_encontrado = False

# NOTAS #NOTAS # NOTAS # Iterar sobre cada iframe y buscar el elemento
for iframe in iframes:
    driver.switch_to.frame(iframe)
    try:
        # NOTAS #NOTAS # NOTAS # Intentar seleccionar el elemento dentro del iframe
        elemento = driver.find_element(By.ID, "form_Ninguno122")
        # NOTAS #NOTAS # NOTAS # Seleccionar el elemento
        elemento.click()
        # NOTAS #NOTAS # NOTAS # El elemento fue encontrado y seleccionado dentro del iframe
        elemento_encontrado = True
        break
    except NoSuchElementException:
        # NOTAS #NOTAS # NOTAS # El elemento no se encontró en este iframe, continuar buscando
        pass
    finally:
        # NOTAS #NOTAS # NOTAS # Volver al contexto principal para buscar en el siguiente iframe
        driver.switch_to.default_content()

# NOTAS #NOTAS # NOTAS # Verificar si se encontró y seleccionó el elemento
if elemento_encontrado:
    print("El elemento se encontró y seleccionó dentro de un iframe.")
else:
    print("El elemento no se encuentra dentro de ningún iframe.")

# NOTAS #NOTAS # NOTAS # Buscar el botón siguiente por su identificador
boton_siguiente = driver.find_element(By.ID, "botonSiguiente")
# NOTAS #NOTAS # NOTAS # Hacer clic en el botón siguiente
boton_siguiente.click()

# NOTAS #NOTAS # NOTAS # Buscar el botón siguiente por su identificador
boton_siguiente = driver.find_element(By.ID, "botonSiguiente")
# NOTAS #NOTAS # NOTAS # Hacer clic en el botón siguiente
boton_siguiente.click()

time.sleep(9)

# NOTAS #NOTAS # NOTAS # Encontrar el botón por tipo y valor
boton_guardar = driver.find_element(
    By.XPATH, '//input[@type="button" and @value="Guardar..."]')

# NOTAS #NOTAS # NOTAS # Obtener el atributo onclick
onclick_attr = boton_guardar.get_attribute("onclick")

# NOTAS #NOTAS # NOTAS # Ejecutar la función webFile.guardarRecibido() utilizando execute_script()
driver.execute_script(onclick_attr)

time.sleep(2)

# NOTAS #NOTAS # NOTAS # Leer el valor del elemento con el id "numRad"
radicado_element = driver.find_element(By.ID, "numRad")
radicado = radicado_element.text

# NOTAS #NOTAS # NOTAS # Obtener la fecha actual
fecha_actual = datetime.datetime.now().strftime("%Y-%m-%d")

# NOTAS #NOTAS # NOTAS # Concatenar la fecha y el valor extraído
valor_completo = f"{fecha_actual} - {radicado}"

# NOTAS #NOTAS # NOTAS # Escribir el valor completo en la columna C del archivo Excel
sheet.cell(row=2, column=3).value = valor_completo

# NOTAS #NOTAS # NOTAS # Obtener la última fila con datos en la columna C
ultima_fila = sheet.max_row

# NOTAS #NOTAS # NOTAS # Obtener los valores de la fila 2
valores_fila_2 = [cell.value for cell in sheet[2]]

# NOTAS #NOTAS # NOTAS # Mover el contenido de la fila 2 a la última fila con espacio libre
fila_destino = ultima_fila + 1
columna_a = sheet.cell(row=2, column=1).value
columna_b = sheet.cell(row=2, column=2).value
columna_c = sheet.cell(row=2, column=3).value

# NOTAS #NOTAS # NOTAS # Agregar una nueva fila al final de la hoja de cálculo
sheet.append([columna_a, columna_b, columna_c] + valores_fila_2)

# NOTAS #NOTAS # NOTAS # Eliminar el contenido de la fila 2
sheet.delete_rows(2)

# NOTAS #NOTAS # NOTAS # Eliminar las columnas D, E y F
sheet.delete_cols(4, amount=3)

# NOTAS #NOTAS # NOTAS # Guardar los cambios en el archivo Excel
workbook.save(excel_file)

# NOTAS #NOTAS # NOTAS # Realizar la combinación de teclas Alt+F4
ActionChains(driver).key_down(Keys.ALT).send_keys(
    Keys.F4).key_up(Keys.ALT).perform()

# NOTAS #NOTAS # NOTAS # Cerrar el navegador
driver.quit()
